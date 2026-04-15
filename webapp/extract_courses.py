"""Extract the Courseware Title Plan sheet into courses.json for the web app."""

import json
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
XLSX = HERE.parent / "titleplan_april10_2026.xlsx"
OUT = HERE / "courses.json"

SHEET = "Courseware Title Plan"

# 1-based column indexes in the sheet.
COL_SOLUTION_AREA = 1
COL_COURSE_NUMBER = 2
COL_TITLE = 3
COL_DURATION = 4
COL_CREDENTIAL = 5
COL_DETAIL_URL = 7


def cell(row, idx):
    value = row[idx - 1]
    if value is None:
        return ""
    return str(value).strip()


def main():
    wb = load_workbook(XLSX, data_only=True, read_only=True)
    ws = wb[SHEET]

    courses = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        course_number = cell(row, COL_COURSE_NUMBER)
        base_url = cell(row, COL_DETAIL_URL)
        if not course_number or not base_url:
            continue
        courses.append({
            "courseNumber": course_number,
            "title": cell(row, COL_TITLE),
            "solutionArea": cell(row, COL_SOLUTION_AREA),
            "duration": cell(row, COL_DURATION),
            "credential": cell(row, COL_CREDENTIAL),
            "baseUrl": base_url,
        })

    OUT.write_text(json.dumps(courses, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {len(courses)} courses to {OUT}")


if __name__ == "__main__":
    main()
